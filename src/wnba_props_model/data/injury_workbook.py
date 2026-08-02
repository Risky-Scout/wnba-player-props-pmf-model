"""Injury workbook loader and leakage validation (audit / label evidence only).

Private workbook rows must never be committed. This module exposes schema,
exact-identity matching, and leakage guards only.
"""
from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable

import pandas as pd

# Event-table columns permitted for ingestion (summary columns excluded).
EVENT_COLUMNS = (
    "number_of__inj",
    "date_injured",
    "athlete",
    "team",
    "body_part",
    "date_returned",
    "total_games_missed",
    "estimated_ws_lost",
    "position",
    "intl./au/college_play?",
)

# Forbidden as onset-time / pre-injury predictive features.
LEAKAGE_PROHIBITED_FEATURES = frozenset({
    "date_returned",
    "total_games_missed",
})

SUMMARY_COLUMN_PREFIXES = (
    "league-wide_games_missed",
    "total_injuries",
    "total_games",
    "games_per_injury",
    "injury_incidence_rate",
    "total_games_rate",
)

YEAR_ONLY_RETURN_SENTINELS = frozenset({"2023", "2024", "2025", "2026"})


@dataclass(frozen=True)
class IdentityMatch:
    athlete: str
    player_id: int | None
    status: str  # exact_roster_name | unresolved | ambiguous_name
    candidates: tuple[int, ...] = ()


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower().replace(" ", "_")
    return text


def is_summary_column(header: str) -> bool:
    h = normalize_header(header)
    return any(h.startswith(p) or h == p for p in SUMMARY_COLUMN_PREFIXES)


def parse_workbook_date(value: Any) -> date | None:
    """Parse injury/return dates; year-only sentinels are not calendar returns."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    if not text or text in YEAR_ONLY_RETURN_SENTINELS:
        return None
    try:
        return pd.to_datetime(text, errors="coerce").date()
    except (TypeError, ValueError):
        return None


def assert_no_onset_leakage(feature_names: Iterable[str]) -> None:
    bad = sorted(set(feature_names) & LEAKAGE_PROHIBITED_FEATURES)
    if bad:
        raise ValueError(
            "injury workbook leakage: prohibited onset-time features: "
            + ", ".join(bad)
        )


def match_athlete_exact(
    athlete: str,
    roster_name_to_ids: dict[str, list[int]],
) -> IdentityMatch:
    """Exact roster-name match only — no unrestricted fuzzy matching."""
    key = " ".join(str(athlete or "").strip().lower().split())
    if not key:
        return IdentityMatch(athlete=athlete, player_id=None, status="unresolved")
    ids = roster_name_to_ids.get(key, [])
    if len(ids) == 1:
        return IdentityMatch(athlete=athlete, player_id=ids[0], status="exact_roster_name")
    if len(ids) > 1:
        return IdentityMatch(
            athlete=athlete,
            player_id=None,
            status="ambiguous_name",
            candidates=tuple(ids),
        )
    return IdentityMatch(athlete=athlete, player_id=None, status="unresolved")


def load_injury_events_from_rows(
    rows: list[dict[str, Any]],
    *,
    roster_name_to_ids: dict[str, list[int]] | None = None,
) -> pd.DataFrame:
    """Normalize event rows (already extracted from event columns only)."""
    roster_name_to_ids = roster_name_to_ids or {}
    out = []
    for r in rows:
        athlete = r.get("athlete")
        match = match_athlete_exact(str(athlete or ""), roster_name_to_ids)
        injured = parse_workbook_date(r.get("date_injured"))
        returned = parse_workbook_date(r.get("date_returned"))
        out.append({
            "athlete": athlete,
            "team": r.get("team"),
            "body_part": r.get("body_part"),
            "date_injured": injured,
            "date_returned": returned,
            "total_games_missed": r.get("total_games_missed"),
            "position": r.get("position"),
            "player_id": match.player_id,
            "identity_status": match.status,
            "return_open": returned is None,
            "season_sheet": r.get("season_sheet"),
        })
    return pd.DataFrame(out)


def load_injury_workbook(
    path: str | Path,
    *,
    roster_name_to_ids: dict[str, list[int]] | None = None,
) -> pd.DataFrame:
    """Load private xlsx using event columns only. Requires openpyxl at runtime."""
    try:
        import openpyxl
    except ImportError as exc:  # pragma: no cover
        raise ImportError("openpyxl required to read injury workbook") from exc

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows: list[dict[str, Any]] = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        it = ws.iter_rows(values_only=True)
        try:
            header = next(it)
        except StopIteration:
            continue
        headers = [normalize_header(h) for h in header]
        # Keep leftmost event block; drop summary columns to the right.
        keep_idx = [
            i for i, h in enumerate(headers)
            if h and not is_summary_column(h) and (
                h in EVENT_COLUMNS or h in {"date_injured", "date_returned", "athlete", "team", "body_part", "total_games_missed", "position"}
            )
        ]
        # Heuristic season from sheet name
        season = None
        for tok in str(sheet).split():
            if tok.isdigit() and len(tok) == 4:
                season = int(tok)
                break
        for raw in it:
            if not raw or all(v is None or str(v).strip() == "" for v in raw):
                continue
            rec = {"season_sheet": season}
            for i in keep_idx:
                if i < len(raw):
                    rec[headers[i]] = raw[i]
            if not rec.get("athlete") or not rec.get("date_injured"):
                continue
            rows.append(rec)
    return load_injury_events_from_rows(rows, roster_name_to_ids=roster_name_to_ids)


def identity_summary(events: pd.DataFrame) -> dict[str, int]:
    if events.empty:
        return {"exact_roster_name": 0, "unresolved": 0, "ambiguous_name": 0}
    vc = events["identity_status"].value_counts().to_dict()
    return {
        "exact_roster_name": int(vc.get("exact_roster_name", 0)),
        "unresolved": int(vc.get("unresolved", 0)),
        "ambiguous_name": int(vc.get("ambiguous_name", 0)),
    }


def eligibility_evidence_from_injury_events(
    events: pd.DataFrame,
    panel: pd.DataFrame,
) -> pd.DataFrame:
    """Build confirmed-inactive eligibility evidence from exact workbook matches.

    Unresolved / ambiguous identities are excluded. ``date_returned`` may close
    an interval for retrospective labeling but must not be emitted as a feature.
    """
    if events.empty or panel.empty:
        return pd.DataFrame()

    exact = events[
        (events["identity_status"] == "exact_roster_name")
        & events["player_id"].notna()
        & events["date_injured"].notna()
    ].copy()
    if exact.empty:
        return pd.DataFrame()

    panel = panel.copy()
    panel["_game_date"] = pd.to_datetime(panel["game_date"], errors="coerce").dt.date
    rows = []
    for _, ev in exact.iterrows():
        pid = int(ev["player_id"])
        start = parse_workbook_date(ev["date_injured"])
        end = parse_workbook_date(ev["date_returned"])  # None => open (esp. 2026)
        if start is None:
            continue
        cand = panel[panel["player_id"] == pid]
        for _, g in cand.iterrows():
            gd = g["_game_date"]
            if gd is None or (isinstance(gd, float) and pd.isna(gd)) or pd.isna(gd):
                continue
            try:
                if gd < start:
                    continue
                if end is not None and gd >= end:
                    continue
            except TypeError:
                continue
            # Player appears on a box row for that game/team context → eligible association
            rows.append({
                "game_id": g["game_id"],
                "player_id": pid,
                "on_eligible_roster": True,
                "injury_interval": True,
                "reviewed_workbook_inactive": True,
                "evidence_timestamp": str(start),
                "label_source": "injury_workbook_exact_identity",
            })
    if not rows:
        return pd.DataFrame()
    return pd.DataFrame(rows).drop_duplicates(["game_id", "player_id"])
